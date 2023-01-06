import openpyxl
from . import detection as dt

class SmartWriter:
  
  HEADERS = [
  "IT",
  "Numero de Expediente",
  "Distrito Judicial", 
  "Instancia", 
  "Especialidad", 
  "Fecha De Inicio",
  "Ano", 
  "N Expedient", 
  "Materia", 
  "Estado",
  "Demandante",
  "Demandado",
  "Seguimiento N#"
  ]

  
  def __init__(self, workbook_name, overwrite) -> None:

    self.workbook_name = workbook_name

    try:
      self.workbook = openpyxl.load_workbook(f"{dt.PROJECT_BASE}/workbooks/{workbook_name}")
      self.overwrite = overwrite
      self.is_new = False
    except:
      self.workbook = openpyxl.Workbook()
      self.workbook.save(f"{dt.PROJECT_BASE}/workbooks/{workbook_name}")
      self.overwrite = True
      self.is_new = True

    if self.overwrite:
      self.clear_file()

    self.queued = []
    self.is_busy = False

  def clear_file(self):
    wb = openpyxl.Workbook()
    wb.save(f"{dt.PROJECT_BASE}/workbooks/{self.workbook_name}")
    self.workbook = openpyxl.load_workbook(f"{dt.PROJECT_BASE}/workbooks/{self.workbook_name}")

  def receive_row(self, row):
    self.queued.append(row)

  def write_row(self, row):
    self.is_busy = True
    ws = self.workbook.active

    def f(data):    
      old_data = self.get_current_file_data()

      if old_data[0][0] == None:
        ws.append(self.HEADERS)
        ws.append(["place","holder"])
        ws.delete_rows(ws.min_row, 1)
        data.insert(0, 1)
        ws.append(data)
      else:
        last_ind = int(old_data[-1][0])
        data.insert(0, last_ind + 1)
        ws.append(data)

      # self.fix_headers()
      # self.sort()
      # self.save()

    f(row)
    for row in self.queued:
      f(row)
    self.queued = []
    self.is_busy = False
  
  def save(self):
    self.workbook.save(f"{dt.PROJECT_BASE}/workbooks/{self.workbook_name}")

  def sort(self):
    print("Writer: Sorting.")
    self.fix_headers(block_print = True)
    data = self.get_current_file_data()
    headers = data.pop(0)
    second_row = data.pop(0)
    
    for line in data:
      line[7] = int(line[7])
      
    data = sorted(data, key= lambda x: x[7])
    
    for i in range(len(data)):
      data[i][0] = i +1 
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(headers)
    new_ws.append(second_row)

    for row in data:
      new_ws.append(row)

    new_wb.save(f"{dt.PROJECT_BASE}/workbooks/{self.workbook_name}")
    self.workbook = new_wb


  def get_current_file_data(self) -> list: 
    ws = self.workbook.active
    
    data = []
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column): 
      vals = []
      for cell in row:
        vals.append(cell.value)
      data.append(vals)
    return data


  def fix_headers(self, block_print = False):
    if not block_print:
      print(f"Writer: Fixing headers.")
    # wb = openpyxl.load_workbook(f"{dt.PROJECT_BASE}/workbooks/{self.workbook_name}")
    og_data = self.get_current_file_data()
    _headers = og_data.pop(0) #pop headers
    _sumilia = og_data.pop(0) #pop sumilia fecha row
    new_og = []
    for data in og_data:
      while data[-1] == None:
        data.pop()
      new_og.append(data)
    og_data = new_og
    sorted_data = sorted(og_data, key= lambda x: len(x), reverse=True)
    
    longest_row = sorted_data[0]

    num_follow_ups = int((len(longest_row) - len(self.HEADERS))/2)
    new_headers = self.HEADERS.copy()
    second_row = []
    new_headers.pop()

    for i in range(num_follow_ups):
      new_headers.append(f"{self.HEADERS[-1]}{i+1}")
      new_headers.append("")
    
    for i in range(len(self.HEADERS)-1):
      second_row.append('')

    for i in range(num_follow_ups):
      second_row.append("SUMILLA")
      second_row.append("FECHA")


    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(new_headers)
    new_ws.append(second_row)

    for row in og_data:
      new_ws.append(row)

    new_wb.save(f"{dt.PROJECT_BASE}/workbooks/{self.workbook_name}")
    self.workbook = new_wb